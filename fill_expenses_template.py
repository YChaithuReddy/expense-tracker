#!/usr/bin/env python3
"""
fill_expenses_template.py

Usage:
  python fill_expenses_template.py \
    --template "/path/to/Expenses Report Format.xlsx" \
    --input-json "/path/to/data.json" \
    --out-dir "./filled"

Input JSON may be a single object or an array of objects (multiple records).
See sample JSON below.

This script edits a copy of the template for each record and preserves styles/formulas.
"""

import argparse
import shutil
import os
import json
import re
from datetime import datetime
from dateutil.parser import parse as dateparse
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

### ---------------------------
### Helper parsing functions
### ---------------------------
def try_parse_date(value):
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        dt = dateparse(str(value), dayfirst=False)  # accept common formats
        return dt.date()
    except Exception:
        # try forcing dayfirst True (e.g., DD/MM/YYYY)
        try:
            dt = dateparse(str(value), dayfirst=True)
            return dt.date()
        except Exception:
            return None

def try_parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    # Remove currency symbols, commas, spaces
    s2 = re.sub(r"[^\d\.\-]", "", s)
    if s2 in ("", "-", ".", "-.", ".-"):
        return None
    try:
        return float(s2)
    except Exception:
        return None

def is_cell_protected(ws, coord):
    """Return True if the cell is protected (worksheet protection + cell locked)."""
    cell = ws[coord]
    # openpyxl stores cell.protection.locked (True by default unless changed)
    # If worksheet is not protected, writing is still possible; but user asked to avoid writing if
    # cell is protected, so we check both sheet protection and cell.locked.
    sheet_protected = ws.protection.sheet
    locked = getattr(cell, "protection", None)
    locked_flag = False
    if locked:
        locked_flag = getattr(cell.protection, "locked", False)
    return sheet_protected and locked_flag

### ---------------------------
### Template mapping (authoritative)
### ---------------------------
HEADER_MAP = {
    "EmployeeName": ("D4", "text"),
    "ExpensePeriod": ("G4", "text"),
    "EmployeeCode": ("D5", "text"),
    "FromDate": ("F5", "date"),
    "ToDate": ("F6", "date"),
    "BusinessPurpose": ("E8", "text"),
}

# Items range: rows 14..66 inclusive -> columns A..F
ITEMS_START_ROW = 14
ITEMS_END_ROW = 66
ITEMS_MAX = ITEMS_END_ROW - ITEMS_START_ROW + 1

CASH_ADV_CELL = "F68"
CHECK_FORMULA_CELLS = ["F67", "F69"]  # expected formulas in the template

SHEET_NAME = "ExpenseReport"  # fallback to first sheet if absent

### ---------------------------
### Main fill routine
### ---------------------------
def fill_one_record(template_path, record, out_dir):
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    base_template_name = os.path.splitext(os.path.basename(template_path))[0]
    out_fname = f"{base_template_name}-filled-{timestamp}.xlsx"
    out_path = os.path.join(out_dir, out_fname)

    # Work on a copy of the template to preserve original
    tmp_copy = out_path + ".tmp.xlsx"
    shutil.copyfile(template_path, tmp_copy)

    wb = load_workbook(tmp_copy)  # keep_vba False by default; if template is xlsm use keep_vba=True and .xlsm path
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]

    validation = []

    # Header fields
    for field, (coord, ftype) in HEADER_MAP.items():
        val = record.get(field)
        if val is None or (isinstance(val, str) and val.strip() == ""):
            validation.append(f"{field}: not provided → left blank")
            continue

        # Check protected
        if is_cell_protected(ws, coord):
            validation.append(f"{field}: target cell {coord} is protected → not written")
            continue

        if ftype == "date":
            dt = try_parse_date(val)
            if dt is None:
                validation.append(f"{field}: value '{val}' unparsable as date → left blank")
            else:
                ws[coord].value = dt
        else:
            ws[coord].value = str(val).strip()

    # CashAdvance
    if "CashAdvance" in record:
        cash = record.get("CashAdvance")
        num = try_parse_number(cash)
        if num is None:
            validation.append(f"CashAdvance: provided '{cash}' not numeric → left blank")
        else:
            if is_cell_protected(ws, CASH_ADV_CELL):
                validation.append(f"CashAdvance: target cell {CASH_ADV_CELL} is protected → not written")
            else:
                ws[CASH_ADV_CELL].value = num
                validation.append(f"CashAdvance: written to {CASH_ADV_CELL} = {num}")

    # Items array
    items = record.get("items", [])
    if not isinstance(items, list):
        validation.append("items: expected array → none written")
        items = []

    ignored = 0
    for idx, item in enumerate(items):
        if idx >= ITEMS_MAX:
            ignored += 1
            continue
        row = ITEMS_START_ROW + idx

        # SLNo -> A
        sl = item.get("SLNo")
        coord_sl = f"A{row}"
        if is_cell_protected(ws, coord_sl):
            validation.append(f"Item {idx+1} SLNo target {coord_sl} protected → not written")
        else:
            if sl is None:
                ws[coord_sl].value = idx + 1
            else:
                try:
                    ws[coord_sl].value = int(sl)
                except Exception:
                    ws[coord_sl].value = idx + 1
                    validation.append(f"Item {idx+1} SLNo '{sl}' non-numeric → auto-numbered {idx+1}")

        # Date -> B
        coord_date = f"B{row}"
        if is_cell_protected(ws, coord_date):
            validation.append(f"Item {idx+1} Date target {coord_date} protected → not written")
        else:
            dt = try_parse_date(item.get("Date"))
            if dt is None:
                ws[coord_date].value = None
                validation.append(f"Item {idx+1} Date '{item.get('Date')}' unparsable → left blank")
            else:
                ws[coord_date].value = dt

        # VendorName_Description -> C
        coord_vendor = f"C{row}"
        if is_cell_protected(ws, coord_vendor):
            validation.append(f"Item {idx+1} Vendor target {coord_vendor} protected → not written")
        else:
            v = item.get("VendorName_Description")
            ws[coord_vendor].value = "" if v is None else str(v).strip()

        # Category -> E
        coord_cat = f"E{row}"
        if is_cell_protected(ws, coord_cat):
            validation.append(f"Item {idx+1} Category target {coord_cat} protected → not written")
        else:
            cat = item.get("Category")
            if cat is None or str(cat).strip() == "":
                ws[coord_cat].value = ""
                validation.append(f"Item {idx+1} Category missing → left blank")
            else:
                ws[coord_cat].value = str(cat).strip()

        # Cost -> F
        coord_cost = f"F{row}"
        if is_cell_protected(ws, coord_cost):
            validation.append(f"Item {idx+1} Cost target {coord_cost} protected → not written")
        else:
            costnum = try_parse_number(item.get("Cost"))
            if costnum is None:
                ws[coord_cost].value = None
                validation.append(f"Item {idx+1} Cost '{item.get('Cost')}' non-numeric → left blank")
            else:
                ws[coord_cost].value = costnum

    if ignored > 0:
        validation.insert(0, f"{ignored} items ignored: only {ITEMS_MAX} rows available (rows {ITEMS_START_ROW}-{ITEMS_END_ROW}).")

    # Verify formula cells remain formulas (soft check)
    for check in CHECK_FORMULA_CELLS:
        c = ws[check]
        # openpyxl stores formula in cell.value as a string beginning with '='; .data_type is sometimes 'f'
        is_formula = isinstance(c.value, str) and str(c.value).startswith("=") or c.data_type == 'f'
        if not is_formula:
            validation.append(f"Warning: {check} expected formula but appears not to contain a formula. Please verify totals.")

    # Save workbook
    # we save to out_path (rename tmp_copy -> out_path)
    wb.save(out_path)

    # remove temp copy
    try:
        os.remove(tmp_copy)
    except Exception:
        pass

    return out_path, validation

### ---------------------------
### CLI and main
### ---------------------------
def main():
    p = argparse.ArgumentParser(description="Fill Expenses Report template while preserving format.")
    p.add_argument("--template", required=True, help="Path to Expenses Report Format.xlsx template")
    p.add_argument("--input-json", required=True, help="Path to JSON input file (single object or array)")
    p.add_argument("--out-dir", default=".", help="Directory to save filled files")
    p.add_argument("--keep-timestamp", action="store_true", help="Keep timestamp in filename (default true)")
    args = p.parse_args()

    template = args.template
    if not os.path.exists(template):
        raise FileNotFoundError(f"Template not found at: {template}")

    with open(args.input_json, "r", encoding="utf-8") as f:
        payload = json.load(f)

    os.makedirs(args.out_dir, exist_ok=True)

    records = payload if isinstance(payload, list) else [payload]

    results = []
    for i, rec in enumerate(records, start=1):
        out_path, validation = fill_one_record(template, rec, args.out_dir)
        results.append((out_path, validation))
        print("=" * 60)
        print(f"Record {i} saved: {out_path}")
        print("Validation Summary:")
        if validation:
            for line in validation:
                print("- " + line)
        else:
            print("- No issues detected.")
        print()

    print("Done. Files saved to:", os.path.abspath(args.out_dir))

if __name__ == "__main__":
    main()